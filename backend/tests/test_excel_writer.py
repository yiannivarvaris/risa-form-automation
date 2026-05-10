from pathlib import Path
import sys

sys.path.append(str(Path(__file__).resolve().parents[2]))

from openpyxl import Workbook, load_workbook

from backend.excel_writer import write_races_to_excel


def _build_template(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "template"
    wb.save(path)


def test_write_races_sorted_and_horse_rows(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    _build_template(template)

    races = [
        {
            "race_number": 2,
            "horses": [
                {"number": 3, "name": "Gamma", "adjusted_current_weight": 55.0, "latest_start": {"rating": 31}},
                {"number": 1, "name": "Alpha", "adjusted_current_weight": 56.0, "latest_start": {"rating": 32}},
            ],
        },
        {
            "race_number": 1,
            "horses": [
                {"number": 2, "name": "Bravo", "adjusted_current_weight": 57.0, "latest_start": {"rating": 29}},
            ],
        },
    ]

    write_races_to_excel(str(template), str(output), races)

    wb = load_workbook(output)
    assert wb.sheetnames == ["template", "Race 1", "Race 2"]

    race_2 = wb["Race 2"]
    assert race_2["A4"].value == 1
    assert race_2["B4"].value == "Alpha"
    assert race_2["A8"].value == 3
    assert race_2["B8"].value == "Gamma"
