import logging

from openpyxl import load_workbook

LOGGER = logging.getLogger(__name__)


def _current_row(horse_number: int) -> int:
    return 4 + ((horse_number - 1) * 2)


def write_races_to_excel(template_path: str, output_path: str, races: list[dict]):
    workbook = load_workbook(template_path)
    if "template" not in workbook.sheetnames:
        raise ValueError("Worksheet 'template' is required")

    template = workbook["template"]

    for sheet_name in list(workbook.sheetnames):
        if sheet_name.startswith("Race "):
            del workbook[sheet_name]

    sorted_races = sorted(races, key=lambda r: int(r.get("race_number", 0)))
    LOGGER.info("sorted race numbers for excel write: %s", [r.get("race_number") for r in sorted_races])

    for index, race in enumerate(sorted_races, start=1):
        race_number = int(race.get("race_number", 0))
        sheet = workbook.copy_worksheet(template)
        sheet.title = f"Race {index}"
        LOGGER.info("created race sheet: %s", sheet.title)

        sorted_horses = sorted(
            race.get("horses", []),
            key=lambda h: int(h.get("number", 10_000)) if str(h.get("number", "")).isdigit() else 10_000,
        )

        for horse in sorted_horses:
            number = horse.get("number")
            if not isinstance(number, int) or number < 1:
                continue

            row = _current_row(number)
            last_row = row + 1
            latest = horse.get("latest_start") or {}
            LOGGER.info(
                "writing horse race=%s number=%s name=%s row=%s latest_start=%s",
                race_number,
                number,
                horse.get("name"),
                row,
                latest.get("date") or latest.get("class") or "none",
            )

            sheet[f"A{row}"] = number
            sheet[f"B{row}"] = horse.get("name")
            sheet[f"P{row}"] = horse.get("adjusted_current_weight")
            sheet[f"U{row}"] = latest.get("class")
            sheet[f"V{row}"] = horse.get("age_sex")

            sheet[f"D{last_row}"] = latest.get("rating", 29)
            sheet[f"E{last_row}"] = latest.get("declared_weight")
            sheet[f"I{last_row}"] = latest.get("actual_carried_weight")
            sheet[f"K{last_row}"] = latest.get("beaten_margin")
            sheet[f"AA{last_row}"] = horse.get("name")
            sheet[f"AD{last_row}"] = horse.get("name")
            sheet[f"AE{last_row}"] = latest.get("runs_since_first_up")
            sheet[f"AF{last_row}"] = latest.get("days_since_last_start")
            sheet[f"AG{last_row}"] = latest.get("track_record")
            sheet[f"AH{last_row}"] = latest.get("distance_record")

            # TODO: template-specific latest start date destination column not provided.
            # Keeping this in a far-right helper column to avoid disrupting existing layout.
            sheet[f"AI{last_row}"] = latest.get("date")

    workbook.save(output_path)
